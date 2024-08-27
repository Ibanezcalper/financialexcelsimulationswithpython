#----------------------------------#
#Autor: Mario Ignacio Ibañez Castro
#Github: https://github.com/MarioIb
#----------------------------------#

#---------------------------------------------------------------------------------------------------------------------------------------#

import openpyxl
import xlsxwriter
import os
import win32com.client
from tqdm import tqdm  # Importar tqdm

# Ruta de la carpeta de salida
output_folder = r'C:\06_Py\05_TRAD03_DOR06\01_salidas\02_SOP'
# Ruta del archivo de resumen
summary_file = r'C:\06_Py\05_TRAD03_DOR06\01_salidas\02_SOP\02_Resumen_SOP_220824.xlsx'

# Crear un nuevo libro para el resumen
workbook_summary = xlsxwriter.Workbook(summary_file)
worksheet_summary = workbook_summary.add_worksheet('Resumen')

# Especificar encabezados
headers = ['Escenario', 'Inversion', 'Tir', 'Tasa de descuento', 'VA Ingresos', 'VA Egresos', 'Costo + Inversion', 'B/C', 'VAN', 'ROI', 'TIIE', 'FC VAN AC 1', 'FC VAN AC 2', 'FC VAN AC 3', 'FC VAN AC 4', 'FC VAN AC 5', 'FC VAN AC 6']
for col_num, header in enumerate(headers):
    worksheet_summary.write(0, col_num, header)

# Función para convertir el índice de columna de Excel (A, B, C, ...) a número (0, 1, 2, ...)
def col_letter_to_num(letter):
    return ord(letter.upper()) - ord('A')

# Función para recalcular y guardar un archivo de Excel utilizando win32com.client
def recalculate_excel(filepath):
    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = False
    wb = excel.Workbooks.Open(filepath)
    wb.RefreshAll()
    wb.Save()
    wb.Close()
    excel.Quit()
dinero_en_anho_1 = 0.7

# Recorrer todos los archivos en la carpeta de salida
row_num = 1

files = [f for f in os.listdir(output_folder) if f.endswith('.xlsx') and f != 'Resumen.xlsx']

# Agregar la barra de progreso
for filename in tqdm(files, desc="Procesando archivos"):
    filepath = os.path.join(output_folder, filename)
    
    # Recalcular y guardar el archivo de Excel
    recalculate_excel(filepath)
    
    # Cargar el libro de trabajo y la hoja deseada
    wb = openpyxl.load_workbook(filepath, data_only=True)
    sheet = wb['Flujo']  # Cambiar al nombre correcto de la hoja si es diferente
    
    predio_name = filename.split('.xlsx')[0]
    values = [predio_name]
    
    # Extraer los valores de las celdas especificadas
    for cell in ['E47','E48', 'E49', 'E50', 'E51', 'E52', 'E53', 'E54', 'E55', 'E56', 'F39', 'F40', 'F41', 'F42', 'F43', 'F44']:
        col = col_letter_to_num(cell[0])
        row = int(cell[1:]) - 1  # Ajustar el índice de fila para openpyxl
        value = sheet.cell(row=row + 1, column=col + 1).value
        values.append(str(value) if value is not None else '')
    
    # Escribir los valores en la hoja de resumen
    for col_num, value in enumerate(values):
        worksheet_summary.write(row_num, col_num, value)
    
    row_num += 1

# Cerrar y guardar el archivo de resumen
workbook_summary.close()
print(f'Se ha guardado el archivo resumen: {summary_file}')